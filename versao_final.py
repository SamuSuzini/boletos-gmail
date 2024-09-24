#%%

import openpyxl as xl
import pdfplumber
import os
import re
import json
from datetime import datetime
from pdfminer.pdfdocument import PDFPasswordIncorrect
from imbox import Imbox
import time

# Funções para extrair dados dos boletos
def extrair_dados_xp(texto):
    # Extrair a primeira data no formato dd/mm/yyyy
    data_vencimento_str = re.search(r'(\d{2}/\d{2}/\d{4})', texto)
    data_vencimento = data_vencimento_str.group(1) if data_vencimento_str else None
    
    # Extrair o valor da fatura
    valor_fatura_str = re.search(r'R\$\s*([\d.,]+)', texto)
    valor_fatura = valor_fatura_str.group(1) if valor_fatura_str else None
    
    # Extrair o código de barras (sequência de 48 números)
    codigo_barras = re.search(r'(\d{5}\.\d{5} \d{5}\.\d{6} \d{5}\.\d{6} \d \d{14})', texto)
    codigo_barras = codigo_barras.group(1) if codigo_barras else None
    
    return data_vencimento, valor_fatura, codigo_barras

def extrair_unimed(texto):
    # Procurar a palavra "ATOS COOPERATIVOS AUXILIARES" seguida pelo valor da fatura
    padrao_valor_fatura = r'ATOS COOPERATIVOS AUXILIARES\s+R\$\s*([\d.,]+)'
    match_valor_fatura = re.search(padrao_valor_fatura, texto, re.IGNORECASE)
    if match_valor_fatura:
        valor_fatura = match_valor_fatura.group(1).replace('.', '').replace(',', '.')
        valor_fatura = "{:.2f}".format(float(valor_fatura)).replace('.', ',')
    else:
        valor_fatura = None
    
    # Extrair a data de vencimento
    data_vencimento_str = re.search(r'(\d{2}/\d{2}/\d{4})', texto)
    data_vencimento = data_vencimento_str.group(1) if data_vencimento_str else None
    
    # Extrair o código de barras (sequência de 48 números)
    codigo_barras = re.search(r'(\d{5}\.\d{5} \d{5}\.\d{6} \d{5}\.\d{6} \d \d{14})', texto)
    codigo_barras = codigo_barras.group(1) if codigo_barras else None
    
    return data_vencimento, valor_fatura, codigo_barras

def extrair_dados_semae(texto):
    # Extrair a data de vencimento
    padrao_data_vencimento = r'(\d{2}/\d{2}/\d{4})'
    match_data_vencimento = re.search(padrao_data_vencimento, texto)
    data_vencimento = match_data_vencimento.group(1) if match_data_vencimento else None

    # Extrair o valor da fatura
    padrao_valor_fatura = r'VALOR ESGOTO\s*(\d+,\d{2})'
    match_valor_fatura = re.search(padrao_valor_fatura, texto)
    valor_fatura = match_valor_fatura.group(1) if match_valor_fatura else None

    # Extrair o código de barras
    padrao_codigo_barras = r'(\d{12} - \d{12} - \d{12} - \d{12})'
    match_codigo_barras = re.search(padrao_codigo_barras, texto)
    codigo_barras = match_codigo_barras.group(1) if match_codigo_barras else None

    return data_vencimento, valor_fatura, codigo_barras

def extrair_dados_nubank(texto):
    # Mapear os meses abreviados para seus equivalentes em números
    meses = {
        "JAN": "01", "FEV": "02", "MAR": "03",
        "ABR": "04", "MAI": "05", "JUN": "06",
        "JUL": "07", "AGO": "08", "SET": "09",
        "OUT": "10", "NOV": "11", "DEZ": "12"
    }
    
    # Extrair a data de vencimento
    padrao_data_vencimento = r'Data\s*do\s*vencimento:\s*(\d{2}\s*[A-Za-z]+\s*\d{4})'
    match_data_vencimento = re.search(padrao_data_vencimento, texto, re.IGNORECASE)
    data_vencimento_str = match_data_vencimento.group(1).strip() if match_data_vencimento else None
    
    if data_vencimento_str:
        # Extrair dia, mês e ano
        partes_data = data_vencimento_str.split()
        dia = partes_data[0]
        mes = meses[partes_data[1].upper()]
        ano = partes_data[2]
        
        # Formatar a data no formato "dd/mm/yyyy"
        data_vencimento = f"{dia}/{mes}/{ano}"
    else:
        data_vencimento = None

    # Extrair o valor da fatura
    padrao_valor_fatura = r'R\$\s*([\d.,]+)'
    match_valor_fatura = re.search(padrao_valor_fatura, texto)
    valor_fatura = match_valor_fatura.group(1).strip() if match_valor_fatura else None
    
    # Extrair o código de barras (sequência de 48 números)
    padrao_codigo_barras = r'(\d{5}\.\d{5} \d{5}\.\d{6} \d{5}\.\d{6} \d \d{14})'
    match_codigo_barras = re.search(padrao_codigo_barras, texto)
    codigo_barras = match_codigo_barras.group(1).strip() if match_codigo_barras else None
    
    return data_vencimento, valor_fatura, codigo_barras


# Define as credenciais e servidores de e-mail
with open("credenciais_gmail.json", "r") as file:
    credenciais = json.loads(file.read())

email = credenciais["e-mail"]
senha = credenciais["senha"]
servidor = credenciais["host"]

# Define o diretório onde os anexos serão salvos
diretorio_anexos = "anexos"

fontes = [
    {"nome": "unimed", "remetente": "digital@unimedriopreto.com.br"},
    {"nome": "semae", "remetente": "conta.semae@sistemas.empro.com.br"},
    {"nome": "cpfl", "remetente": "contadigital@cpfl.com.br"},
    {"nome": "xp", "remetente": "fatura@xpi.com.br"},
    {"nome": "nubank", "remetente": "todomundo@nubank.com.br"}
]

diretorio_anexos = "anexos"
if not os.path.exists(diretorio_anexos):
    os.makedirs(diretorio_anexos)

with Imbox(
    hostname=servidor,
    username=email,
    password=senha) as imbox:

    # Lista para armazenar todas as mensagens
    todas_mensagens = []

    # Itera sobre as fontes de e-mail
    for fonte in fontes:
        # Busca as mensagens da caixa de entrada para o remetente especificado
        mensagens = imbox.messages(sent_from=fonte["remetente"])
        # Adiciona as mensagens da fonte atual à lista de todas as mensagens
        todas_mensagens.extend(mensagens)

    # Itera sobre todas as mensagens coletadas
    for uid, msg in todas_mensagens:
        remetente = msg.sent_from[0]['email']
        for fonte in fontes:
            if remetente == fonte["remetente"]:
                # print(f"Mensagem da conta {fonte['nome']}:")
                # print(f"De: {remetente}")
                # print(f"Assunto: {msg.subject}")
                
                # Itera sobre os anexos da mensagem e os salva
                for anexo in msg.attachments:
                    nome_arquivo = anexo["filename"]
                    conteudo = anexo["content"]
                    
                    # Salva o anexo no diretório com o nome do remetente adicionado ao nome do arquivo
                    caminho_arquivo = os.path.join(diretorio_anexos, f"{fonte['nome']}-{nome_arquivo}")
                    with open(caminho_arquivo, "wb") as file:
                        file.write(conteudo.read())

        caixa_entrada = imbox.messages()
        qtde_nao_lidos = imbox.messages(unread=True)
        qtde_lidos = (len(caixa_entrada)) - (len(qtde_nao_lidos))
        qtde_nao_lidos = (len(qtde_nao_lidos))

        

# Carregar a planilha do Excel
excel = xl.load_workbook('BD-Boletos.xlsx')
aba = excel.active
aba.title = 'BD'
excel.save('BD-Boletos.xlsx')

# Carregar o arquivo JSON com as senhas
with open('credenciais_gmail.json', 'r') as f:
    credenciais = json.load(f)
senhas_pdf = credenciais.get("senhas_pdf", {})


time.sleep(3)
# Diretório com os arquivos PDF
diretorio_pdfs = 'anexos'


# Iterar sobre os arquivos PDF no diretório
for arquivo in os.listdir(diretorio_pdfs):
    if arquivo.endswith('.pdf'):
       
        # Caminho completo do arquivo PDF
        arquivo_pdf = os.path.join(diretorio_pdfs, arquivo)
        
        # Nome do arquivo PDF e beneficiário
        nome_arquivo_pdf = os.path.basename(arquivo_pdf)
        beneficiario = nome_arquivo_pdf.split('-')[0].upper()
        
        # Verificar se há senha para o beneficiário
        senha = senhas_pdf.get(beneficiario.lower())
        
        try:
            # Abrir o PDF e extrair o texto da primeira página
            if senha:
                pdf = pdfplumber.open(arquivo_pdf, password=senha)
            else:
                pdf = pdfplumber.open(arquivo_pdf)
                
            pagina = pdf.pages[0]
            texto_pag1 = pagina.extract_text()
            pdf.close()

            # Extrair a data de vencimento, valor da fatura e código de barras
            if beneficiario == "UNIMED":
                data_vencimento, valor_fatura, codigo_barras = extrair_unimed(texto_pag1)
            elif beneficiario == "SEMAE":
                data_vencimento, valor_fatura, codigo_barras = extrair_dados_semae(texto_pag1)
            elif beneficiario == "NUBANK":
                data_vencimento, valor_fatura, codigo_barras = extrair_dados_nubank(texto_pag1)
            else:
                # Caso não seja um beneficiário conhecido, utilizar a função padrão (XP)
                data_vencimento, valor_fatura, codigo_barras = extrair_dados_xp(texto_pag1)

            # print(f"Beneficiário: {beneficiario}")
            # print(f"Data de Vencimento: {data_vencimento}")
            # print(f"Valor da Fatura: {valor_fatura}")
            # print(f"Código de Barras: {codigo_barras}")

            # Determinar a próxima linha vazia na planilha
            row_inicio = len(aba["A"]) + 1
            data_extracao = datetime.now().strftime('%d/%m/%Y')

            # Escrever os dados na planilha
            aba.cell(row=row_inicio, column=1).value = data_extracao
            aba.cell(row=row_inicio, column=2).value = beneficiario
            aba.cell(row=row_inicio, column=3).value = data_vencimento
            aba.cell(row=row_inicio, column=4).value = codigo_barras
            aba.cell(row=row_inicio, column=5).value = valor_fatura

            # Salvar a planilha
            
            excel.save('BD-Boletos.xlsx')

        except PDFPasswordIncorrect:
            print(f'O arquivo {arquivo_pdf} está protegido por senha e não foi possível abrir com a senha fornecida.')


print(f"Você tem {qtde_nao_lidos} não lidos e {qtde_lidos} lidos em sua caixa de entrada!")

#%%